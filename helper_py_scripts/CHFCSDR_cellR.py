#!/usr/bin/env python

# Solo didn't run through scvi, scvi-tools nor scanpy.external
# Only this seems to work
from solo import hashsolo
import anndata as ad
import scanpy as sc, pandas as pd, numpy as np
import glob2, os, re
from collections import Counter
from openpyxl import load_workbook
from collections import defaultdict
import subprocess
from pathlib import Path 


sc.settings.set_figure_params(dpi_save=400, format='png', color_map = 'viridis_r')
sc.settings.autosave = True
sc.settings.autoshow = False
sc.settings.verbosity = 3  # verbosity: errors (0), warnings (1), info (2), hints (3)
sc.logging.print_version_and_date()


files = [f for f in glob2.glob('/sc/arion/projects/psychAD/Single_cell_data/alignment/*/cDNA/results/*/outs/filtered_feature_bc_matrix/matrix.mtx.gz')]

#def def_value():
    #return "Not Present"


max_mito=5
min_genes=1000
min_cells=10


t2g = pd.read_csv("/sc/arion/projects/psychAD/pnm/Hs_allchr_MT.txt", skiprows=1, usecols=range(2),names=["gene_id", "gene_name"], sep="\t")
t2g.index = t2g.gene_id
t2g = t2g.loc[~t2g.index.duplicated(keep='first')]

# Calculate how many cells weren't found in solo_hashed out
def ret_htos(bcs, df_shan, b):
    # List of (barcode, HTO)
   barc_l = []
   #hto_l = []
   #samp_n = []
   ret_dict = {}
   doublet_n = 0
   negative_n = 0
   #print(df_shan.columns)
   for bc in bcs:
       if bc in hash_data.obs_names:
           #hto_l.append(hash_data.obs[hash_data.obs_names == bc].Classification.values[0])
           hto_n = hash_data.obs[hash_data.obs_names == bc].Classification.values[0]
           #samp_n.append(df_shan[(df_shan["cDNA_ID"] == b) & (df_shan["hashtag" == hto_n]), "sample_ID"].values[0])

           if hto_n == 'Doublet':
               doublet_n += 1

           elif hto_n == 'Negative':
               negative_n += 1

           else:
               barc_l.append(bc)
               #print(b)
               #print(hto_n)
               #print(df_shan.loc[(df_shan["cDNA_ID"] == b) & (df_shan["hashtag"] == hto_n), "sample_ID"].values[0])
               try:
                   ret_dict[bc] = df_shan.loc[(df_shan["cDNA_ID"] == b) & (df_shan["hashtag"] == hto_n), "SubID"].values[0]
               except:
                   print(b)
                   #print(df_shan["cDNA_ID"].head())
                   print( df_shan.loc[(df_shan["cDNA_ID"] == b)].head())
                   print('\n\n\n')
                   print(df_shan.loc[(df_shan["cDNA_ID"] == b) & (df_shan["hashtag"] == hto_n)].head())
                   #print(f'For the Sample: {b} and for the HTO tag {hto_n} there was no data in the excel file!')



   return [barc_l, ret_dict, doublet_n, negative_n] #zip(barc_l, samp_n)



def ret_samp_names(y, bc_dict):
   return bc_dict[y]



# Download Shan's spreadsheet if already not present
try:
   wb = load_workbook('wet_lab_data_shan.xlsx')
except:
   get_file = subprocess.run(["wget", "--no-check-certificate", "https://docs.google.com/spreadsheets/d/1t5Ed3X_FOFB0_C9GFG_TEW9XDZzJcHTZHiP26HN2yxQ/export?format=xlsx", "-O", "wet_lab_data_shan.xlsx"])
   wb = load_workbook('wet_lab_data_shan.xlsx')

df_l = []
for row in wb[wb.sheetnames[0]].iter_rows(min_col=2, max_col=47, min_row=1, values_only=True):
   df_l.append([re.sub('[-/\(\)]', '_', str(row[0])).replace(' ', ''), row[9], row[44].replace('#', 'HTO'), row[45]])


headers=df_l.pop(0)
headers[0] = headers[0].replace('.', '_')
headers[3] = 'barcode'
# Columns: SubID, cDNA_ID, hashtag, barcode
df = pd.DataFrame(df_l, columns=headers)



for inp_f in files:
   
   if Path().is_file():
       continue

   else:

       # Batch and round info
       #batch=re.search('Sample-(NPSAD-.*)\.h5ad', os.path.basename(snakemake.output[0])).group(1)
       #r_num=os.path.basename(snakemake.output[0]).split('_')[0].replace('round', '')
       batch=re.search('/Sample_(.*-cDNA)/', inp_f).group(1)
       r_num=re.search('/round([0-9]+)/', inp_f).group(1)
   
       #hash_data = ad.read(snakemake.input[0])
       hash_data = ad.read("/sc/arion/projects/psychAD/pnm/cache/round{}_Sample-{}.h5ad".format(r_num, batch))
       hashsolo.hashsolo(hash_data)
       '''

       Default:
       priors: list = [.01, .8, .19],
       pre_existing_clusters: str = None,
       clustering_data: anndata.AnnData = None,
       resolutions: list = [.1, .25, .5, .75, 1],
       number_of_noise_barcodes: int = None,
       inplace: bool = True

       '''

       # Hashing output
       #hash_data.write(snakemake.output[0])
       hash_data.write("/sc/arion/projects/psychAD/demultiplex/solo/round{}_Sample-{}_cellR_calicosolo_out.h5ad".format(r_num, batch))


       # Parameters for filtering
       #max_mito=snakemake.params[0]
       #min_genes=snakemake.params[1]
       #min_cells=snakemake.params[2]


       #adata = sc.read_10x_mtx(snakemake.input[1][:-13], make_unique=True, var_names= "gene_ids", cache=True)
       adata = sc.read_10x_mtx(inp_f[:-13], make_unique=True, var_names= "gene_ids", cache=True)
       solo_run_info = []
       solo_run_info.append(( 'Started with cells', adata.n_obs))
       solo_run_info.append(( 'Started with genes', adata.n_vars))
       adata.var_names_make_unique()
       adata.var["gene_id"] = adata.var.index.values
       adata.var["gene_name"] = adata.var.gene_id.map(t2g["gene_name"])
       adata.var_names = adata.var_names.to_series().map(lambda x: x + '_index')
       adata = adata[:, pd.notna(adata.var["gene_name"])]
       adata.var_names_make_unique()
       adata.X = adata.X.astype('float64')
       solo_run_info.append(( 'Unique genes', adata.n_vars))
    
       # Filter data using cell level metrics
       sc.pp.filter_cells(adata, min_genes=min_genes)
       sc.pp.filter_genes(adata, min_cells=min_cells)
       solo_run_info.append(( 'min #genes expressed per cell', min_genes))
       solo_run_info.append(( 'min #cells expressing per gene', min_cells))
       solo_run_info.append(( 'Retained cells after previous filter', adata.n_obs))
       solo_run_info.append(( 'Retained genes after previous filter', adata.n_vars))
   
       # Filter data wrt mito content
       adata.var["mito"] = adata.var_names.str.startswith('MT-')
       sc.pp.calculate_qc_metrics(adata, inplace=True, qc_vars=["mito"])
       adata = adata[adata.obs["pct_counts_mito"]< 1, :]
       solo_run_info.append(( 'max percent mito content per cell', max_mito))
       solo_run_info.append(( 'Started with cells', adata.n_obs))
       solo_run_info.append(( 'Started with genes', adata.n_vars))

       # Demultiplex and assign samples (after downloading the spreadsheet)
       hto_tags = ret_htos(adata.obs_names.tolist(), df, batch)
       #barc_list = []
       #samp_name_list = []

       adata = adata[hto_tags[0]].copy()
       adata.obs['samp_name'] = adata.obs_names.to_series().apply(ret_samp_names, args=(hto_tags[1],))

       solo_run_info.append(('Doublets #cells', hto_tags[2]))
       solo_run_info.append(('Negative #cells', hto_tags[3]))
       adata.obs['batch'] = batch
       adata.obs['round_num'] = r_num

       solo_run_info.append(( 'After demultiplexing #cells', adata.n_obs))


       #adata.obs['sample_name'] = ret_samp_names(adata.obs_names.tolist(), barc_list, HTO_list, df)


       solo_run_df = pd.DataFrame(solo_run_info, columns=['Observations', 'Vals'])
       solo_run_df.to_csv("/sc/arion/projects/psychAD/demultiplex/solo/{}_Sample-{}_cellR_calicosolo_info.tsv".format(r_num, batch), sep = "\t", index=False)
       #solo_run_df.to_csv(snakemake.output[2], sep = "\t", index=False)

       adata.write("/sc/arion/projects/psychAD/final_count_matrix/solo/{}_Sample-{}_cellR_calicosolo_out.h5ad".format(r_num, batch))
       #ad.write(snakemake.output[1])
