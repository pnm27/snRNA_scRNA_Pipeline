#!/sc/arion/work/prashf01/conda/envs/snakemake/bin/python

# Solo didn't run through scvi, scvi-tools nor scanpy.external
# Only this seems to work
import anndata as ad
import scanpy as sc, pandas as pd, numpy as np
import glob2, os, re
from collections import Counter
from openpyxl import load_workbook
from collections import defaultdict
# using datetime module
import datetime



sc.settings.set_figure_params(dpi_save=400, format='png', color_map = 'viridis_r')
sc.settings.autosave = True
sc.settings.autoshow = False
sc.settings.verbosity = 3  # verbosity: errors (0), warnings (1), info (2), hints (3)
sc.logging.print_version_and_date()

hashsolo_out = "/sc/arion/projects/psychAD/demultiplex/solo/round2_Sample-NPSAD-20201125-A2-cDNA_calicosolo_out.h5ad"
multiseq_out = "/sc/arion/projects/psychAD/demultiplex/MULTIseq/round2_Sample-NPSAD-20201125-A2-cDNA_MUTLIseq_out.h5ad"
cellR_mat = "/sc/arion/projects/psychAD/Single_cell_data/alignment/round2/cDNA/results/Sample_NPSAD-20201125-A2-cDNA/outs/filtered_feature_bc_matrix/matrix.mtx.gz"[:-13]
starsolo_mat = "/sc/arion/projects/psychAD/STARsolo_bams/round2/Sample_NPSAD-20201125-A2-cDNA/NPSAD-20201125-A2-cDNA_L001_001_Solo.out/GeneFull/filtered_Lun/matrix.mtx.gz"[:-13]

# Batch and round info
batch=re.search('Sample_(NPSAD-.*)/outs/', cellR_mat).group(1)
r_num=2
preparer=batch.split('-')[2]
replicate=preparer[1]

# Parameters for filtering
max_mito=5
min_genes=1000
min_cells=10


dem_cs = ad.read(hashsolo_out)
dem_ms = ad.read(multiseq_out)

t2g = pd.read_csv("/sc/arion/projects/psychAD/pnm/Hs_allchr_MT.txt", skiprows=1, usecols=range(2),names=["gene_id", "gene_name"], sep="\t")
t2g.index = t2g.gene_id
t2g = t2g.loc[~t2g.index.duplicated(keep='first')]

# Calculate how many cells weren't found in solo_hashed out
def ret_htos_calico_solo(bcs, df_shan, b):
    # List of (barcode, HTO)
   barc_l = []
   #hto_l = []
   #samp_n = []
   ret_dict = {}
   doublet_n = 0
   negative_n = 0
   for bc in bcs:
       if bc in dem_cs.obs_names:
           #hto_l.append(hash_data.obs[hash_data.obs_names == bc].Classification.values[0])
           hto_n = dem_cs.obs[dem_cs.obs_names == bc].Classification.values[0]
           #samp_n.append(df_shan[(df_shan["cDNA_ID"] == b) & (df_shan["hashtag" == hto_n]), "sample_ID"].values[0])
           if hto_n == 'Doublet':
               doublet_n += 1

           elif hto_n == 'Negative':
               negative_n += 1

           else:
               barc_l.append(bc)
               ret_dict[bc] = df_shan.loc[(df_shan["cDNA_ID"] == b) & (df_shan["hashtag"] == hto_n), "SubID"].values[0]
               #try:
                   #ret_dict[bc] = df_shan.loc[(df_shan["cDNA_ID"] == b) & (df_shan["hashtag"] == hto_n), "SubID"].values[0]
               #except:
                   #if 'NPSAD-20201021-A1-cDNA' != b or 'NPSAD-20201021-A2-cDNA' != b or 'NPSAD-20201022-A1-cDNA' != b or 'NPSAD-20201022-A2-cDNA' != b:
                       #print(f'For the Sample: {b} and for the HTO tag {hto_n} there was no data in the excel file!')


   return [barc_l, ret_dict, doublet_n, negative_n] #zip(barc_l, samp_n)



def ret_htos_multiseq(bcs, df_shan, b):
    # List of (barcode, HTO)
   barc_l_m = []
   #hto_l = []
   #samp_n = []
   ret_dict_m = {}
   m_doublet_n = 0
   m_negative_n = 0
   unc = 0
   for bc in bcs:
       if bc in dem_ms.obs_names:
           #hto_l.append(hash_data.obs[hash_data.obs_names == bc].Classification.values[0])
           hto_n = dem_ms.X[dem_ms.obs_names == bc].toarray().sum(axis=1)
           #samp_n.append(df_shan[(df_shan["cDNA_ID"] == b) & (df_shan["hashtag" == hto_n]), "sample_ID"].values[0])
           if hto_n > 1:
               m_doublet_n += 1

           elif hto_n == 0:
               m_negative_n += 1

           elif hto_n == 1:
               barc_l_m.append(bc)
               row, col = np.where(dem_ms.X[dem_ms.obs_names == bc].toarray() == 1)
               hto_num = dem_ms.var_names[col].values[0]
               ret_dict[bc] = df_shan.loc[(df_shan["cDNA_ID"] == b) & (df_shan["hashtag"] == hto_num), "SubID"].values[0]
               #try:
                   #ret_dict[bc] = df_shan.loc[(df_shan["cDNA_ID"] == b) & (df_shan["hashtag"] == hto_n), "SubID"].values[0]
               #except:
                   #if 'NPSAD-20201021-A1-cDNA' != b or 'NPSAD-20201021-A2-cDNA' != b or 'NPSAD-20201022-A1-cDNA' != b or 'NPSAD-20201022-A2-cDNA' != b:
                       #print(f'For the Sample: {b} and for the HTO tag {hto_n} there was no data in the excel file!')
           else:
               unc += 1

    
   print(f"Unclassified # of cells: {unc}")

   return [barc_l_m, ret_dict_m, m_doublet_n, m_negative_n] #zip(barc_l, samp_n)



def ret_samp_names(y, bc_dict):
   if y in bc_dict.keys():
       return bc_dict[y]
   
   else:
       return "Not Present"



def ret_bad_cells(cell, cs_bc, ms_bc):
   if cell in cs_bc or cell in ms_bc:
       return cell



# Shan's new csv file
df = pd.read_csv('/sc/arion/projects/psychAD/upload2synpase/final_NPSAD_snRNAseq_metrics_combo.csv')

# Process STARsolo output----------------------------------------------------------------------------------------------------------------------------
ct = datetime.datetime.now()
print(f"Processing STARsolo's output at: {ct}")

adata = sc.read_10x_mtx(starsolo_mat, make_unique=True, var_names= "gene_ids", cache=True)
solo_run_info = []
solo_run_info.append(( 'Started with cells', adata.n_obs))
solo_run_info.append(( 'Started with genes', adata.n_vars))
adata.var_names_make_unique()
adata.var["gene_id"] = adata.var.index.values
adata.var["gene_name"] = adata.var.gene_id.map(t2g["gene_name"])
adata.var_names = adata.var_names.to_series().map(lambda x: x + '_index')
adata = adata[:, pd.notna(adata.var["gene_name"])]
adata.var_names_make_unique()
adata.X = adata.X.astype('float64')
solo_run_info.append(( 'Unique genes', adata.n_vars))

# Filter data using cell level metrics
sc.pp.filter_cells(adata, min_genes=min_genes)
sc.pp.filter_genes(adata, min_cells=min_cells)
solo_run_info.append(( 'min #genes expressed per cell', min_genes))
solo_run_info.append(( 'min #cells expressing per gene', min_cells))
solo_run_info.append(( 'Retained cells after previous filter', adata.n_obs))
solo_run_info.append(( 'Retained genes after previous filter', adata.n_vars))

# Filter data wrt mito content
adata.var["mito"] = adata.var_names.str.startswith('MT-')
sc.pp.calculate_qc_metrics(adata, inplace=True, qc_vars=["mito"])
adata = adata[adata.obs["pct_counts_mito"]< max_mito, :]
solo_run_info.append(( 'max percent mito content per cell', max_mito))
solo_run_info.append(( 'cells with low mito percent', adata.n_obs))

# Download Shan's spreadsheet if already not present
#try:
   #wb = load_workbook('wet_lab_data_shan.xlsx')
#except:
   #get_file = subprocess.run(["wget", "--no-check-certificate", "https://docs.google.com/spreadsheets/d/10HYHDlsvtmBK_j3AgfPCnvaNohAg-OBhXi0iOzR9e0A/export?format=xlsx", "-O", "wet_lab_data_shan.xlsx"])
   #wb = load_workbook('wet_lab_data_shan.xlsx')

#df_l = []
#for row in wb[wb.sheetnames[0]].iter_rows(min_col=2, max_col=47, min_row=1, values_only=True):
   #df_l.append([re.sub('[-/\(\)]', '_', str(row[0])).replace(' ', ''), row[9], row[44].replace('#', 'HTO'), row[45]])


#headers=df_l.pop(0)
#headers[0] = headers[0].replace('.', '_')
#headers[3] = 'barcode'
# Columns: Sub_ID, cDNA_ID, hashtag, barcode
#df = pd.DataFrame(df_l, columns=headers)

# Demultiplex and assign samples (after downloading the spreadsheet)---------------------------------------------

# hto_tags_cs and hto_tags_ms contain (in sequence): barcode, dictionary with barcodes as key and SubID as value
#, no. of doublet cells, no. of negative cells]
ct = datetime.datetime.now()
print(f"Starting: Calculating demultiplexing info for hashsolo/calico solo at: {ct}")
hto_tags_cs = ret_htos_calico_solo(adata.obs_names.to_list(), df, batch)
ct = datetime.datetime.now()
print(f"Finished: Calculating demultiplexing info for hashsolo/calico solo at: {ct}")
print(f"Starting: Calculating demultiplexing info for MULTIseq at: {ct}")
hto_tags_ms = ret_htos_multiseq(adata.obs_names.to_list(), df, batch)
ct = datetime.datetime.now()
print(f"Finished: Calculating demultiplexing info for MULTIseq at: {ct}")

#barc_list = []
#samp_name_list = []

common_bc, cs_ind, ms_ind = np.intersect1d(hto_tags_cs[0], hto_tags_ms[0], return_indices=True)
SubID_cs = adata.obs_names.to_series().apply(ret_samp_names, args=(hto_tags_cs[1],))
adata.obs['SubID_cs'] = SubID_cs
# Save output from calico_solo
adata[hto_tags_cs[0]].copy().write('test_ss_cs.h5ad')

adata.obs['SubID_ms'] = adata.obs_names.to_series().apply(ret_samp_names, args=(hto_tags_ms[1],))
del adata.obs['SubID_cs'] # Don't want this obs in other prog's out
# Save output from MULTIseq
adata[hto_tags_ms[0]].copy().write('test_ss_ms.h5ad')

# for anno in either
adata.obs['SubID_cs'] = SubID_cs


solo_run_info.append(('Doublets #cells_cs', hto_tags_cs[2]))
solo_run_info.append(('Negative #cells_cs', hto_tags_cs[3]))
solo_run_info.append(('Doublets #cells_ms', hto_tags_ms[2]))
solo_run_info.append(('Negative #cells_ms', hto_tags_ms[3]))

rem_cells_cs = adata.obs.SubID_cs.value_counts()
rem_cells_ms = adata.obs.SubID_ms.value_counts()
solo_run_info.append(( 'After demultiplexing #cells_cs', rem_cells_cs[rem_cells_cs.index != "Not Present"] ))
solo_run_info.append(( 'After demultiplexing #cells_ms', rem_cells_ms[rem_cells_ms.index != "Not Present"] ))

# bc present in any demux method:
good_bcs = adata.obs_names.to_series().apply(ret_bad_cells, args=(hto_tags_cs[0], hto_tags_ms[0])).to_list()
good_bcs = [x for x in good_bcs if x is not None]

ct = datetime.datetime.now()
print(f"Subsetting dataset to retain only cells annotated (can be different) by both dem methods at: {ct}")

# Subset only the previous cells
adata = adata[good_bcs].copy()

# add few more annotations
adata.obs['batch'] = batch
adata.obs['round_num'] = r_num
adata.obs['prep'] = preparer
adata.obs['rep'] = replicate

solo_run_df = pd.DataFrame(solo_run_info, columns=['Observations', 'Vals'])
solo_run_df.to_csv('test_ss.tsv', sep = "\t", index=False)

adata.write('test_ss.h5ad')

#----------------------------------------------------------------------------------------------------------------------------------------------

# Process cellranger output----------------------------------------------------------------------------------------------------------------------------
ct = datetime.datetime.now()
print(f"Starting to process ccellranger dataset at: {ct}")
adata_cr = sc.read_10x_mtx(cellR_mat, make_unique=True, var_names= "gene_ids", cache=True)
solo_run_info = []
solo_run_info.append(( 'Started with cells', adata_cr.n_obs))
solo_run_info.append(( 'Started with genes', adata_cr.n_vars))
adata_cr.var_names_make_unique()
adata_cr.var["gene_id"] = adata_cr.var.index.values
adata_cr.var["gene_name"] = adata_cr.var.gene_id.map(t2g["gene_name"])
adata_cr.var_names = adata_cr.var_names.to_series().map(lambda x: x + '_index')
adata_cr = adata_cr[:, pd.notna(adata_cr.var["gene_name"])]
adata_cr.var_names_make_unique()
adata_cr.X = adata_cr.X.astype('float64')
solo_run_info.append(( 'Unique genes', adata_cr.n_vars))

# Filter data using cell level metrics
sc.pp.filter_cells(adata_cr, min_genes=min_genes)
sc.pp.filter_genes(adata_cr, min_cells=min_cells)
solo_run_info.append(( 'min #genes expressed per cell', min_genes))
solo_run_info.append(( 'min #cells expressing per gene', min_cells))
solo_run_info.append(( 'Retained cells after previous filter', adata_cr.n_obs))
solo_run_info.append(( 'Retained genes after previous filter', adata_cr.n_vars))

# Filter data wrt mito content
adata_cr.var["mito"] = adata_cr.var_names.str.startswith('MT-')
sc.pp.calculate_qc_metrics(adata_cr, inplace=True, qc_vars=["mito"])
adata_cr = adata_cr[adata_cr.obs["pct_counts_mito"]< max_mito, :]
solo_run_info.append(( 'max percent mito content per cell', max_mito))
solo_run_info.append(( 'cells with low mito percent', adata_cr.n_obs))


ct = datetime.datetime.now()
print(f"Starting: Calculating demultiplexing info for hashsolo/calico solo at: {ct}")
hto_tags_cs = ret_htos_calico_solo(adata_cr.obs_names.to_list(), df, batch)
ct = datetime.datetime.now()
print(f"Finished: Calculating demultiplexing info for hashsolo/calico solo at: {ct}")
print(f"Starting: Calculating demultiplexing info for MULTIseq at: {ct}")
hto_tags_ms = ret_htos_multiseq(adata_cr.obs_names.to_list(), df, batch)
ct = datetime.datetime.now()
print(f"Finished: Calculating demultiplexing info for MULTIseq at: {ct}")

common_bc, cs_ind, ms_ind = np.intersect1d(hto_tags_cs[0], hto_tags_ms[0], return_indices=True)
SubID_cs = adata_cr.obs_names.to_series().apply(ret_samp_names, args=(hto_tags_cs[1],))
adata_cr.obs['SubID_cs'] = SubID_cs
# Save output from calico_solo
adata_cr[hto_tags_cs[0]].copy().write('test_cr_cs.h5ad')

adata_cr.obs['SubID_ms'] = adata_cr.obs_names.to_series().apply(ret_samp_names, args=(hto_tags_ms[1],))
del adata_cr.obs['SubID_cs'] # Don't want this obs in other prog's out
# Save output from MULTIseq
adata_cr[hto_tags_ms[0]].copy().write('test_cr_ms.h5ad')

# for anno in either
adata_cr.obs['SubID_cs'] = SubID_cs


solo_run_info.append(('Doublets #cells_cs', hto_tags_cs[2]))
solo_run_info.append(('Negative #cells_cs', hto_tags_cs[3]))
solo_run_info.append(('Doublets #cells_ms', hto_tags_ms[2]))
solo_run_info.append(('Negative #cells_ms', hto_tags_ms[3]))

rem_cells_cs = adata_cr.obs.SubID_cs.value_counts()
rem_cells_ms = adata_cr.obs.SubID_ms.value_counts()
solo_run_info.append(( 'After demultiplexing #cells_cs', rem_cells_cs[rem_cells_cs.index != "Not Present"] ))
solo_run_info.append(( 'After demultiplexing #cells_ms', rem_cells_ms[rem_cells_ms.index != "Not Present"] ))

# bc present in any demux method:
good_bcs = adata_cr.obs_names.to_series().apply(ret_bad_cells, args=(hto_tags_cs[0], hto_tags_ms[0])).to_list()
good_bcs = [x for x in good_bcs if x is not None]

ct = datetime.datetime.now()
print(f"Subsetting dataset to retain only cells annotated (can be different) by both dem methods at: {ct}")


# Subset only the previous cells
adata_cr = adata_cr[good_bcs].copy()

# add few more annotations
adata_cr.obs['batch'] = batch
adata_cr.obs['round_num'] = r_num
adata_cr.obs['prep'] = preparer
adata_cr.obs['rep'] = replicate

solo_run_df = pd.DataFrame(solo_run_info, columns=['Observations', 'Vals'])
solo_run_df.to_csv('test_cr.tsv', sep = "\t", index=False)

adata_cr.write('test_cr.h5ad')

ct = datetime.datetime.now()
print(f"Finished: Processing Sample {batch} at: {ct}")
