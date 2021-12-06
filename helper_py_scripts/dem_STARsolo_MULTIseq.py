#!/sc/arion/work/prashf01/conda/envs/snakemake/bin/python

# Solo didn't run through scvi, scvi-tools nor scanpy.external
# Only this seems to work
from solo import hashsolo
import anndata as ad
import scanpy as sc, pandas as pd, numpy as np
import glob2, os, re
from collections import Counter
from openpyxl import load_workbook
from collections import defaultdict


sc.settings.set_figure_params(dpi_save=400, format='png', color_map = 'viridis_r')
sc.settings.autosave = True
sc.settings.autoshow = False
sc.settings.verbosity = 3  # verbosity: errors (0), warnings (1), info (2), hints (3)
sc.logging.print_version_and_date()

#print(os.environ['PATH'])

#def def_value():
    #return "Not Present"



hash_data = ad.read(snakemake.input[0])
hashsolo.hashsolo(hash_data)
'''

Default:
priors: list = [.01, .8, .19],
pre_existing_clusters: str = None,
clustering_data: anndata.AnnData = None,
resolutions: list = [.1, .25, .5, .75, 1],
number_of_noise_barcodes: int = None,
inplace: bool = True

'''

# Hashing output
hash_data.write(snakemake.output[0])

# Batch and round info
batch=re.search('Sample-(NPSAD-.*)_calicosolo_out.h5ad', os.path.basename(snakemake.output[0])).group(1)
r_num=int(os.path.basename(snakemake.output[0]).split('_')[0].replace('round', ''))

# Parameters for filtering
max_mito=snakemake.params[0]
min_genes=snakemake.params[1]
min_cells=snakemake.params[2]


t2g = pd.read_csv("/sc/arion/projects/psychAD/pnm/Hs_allchr_MT.txt", skiprows=1, usecols=range(2),names=["gene_id", "gene_name"], sep="\t")
t2g.index = t2g.gene_id
t2g = t2g.loc[~t2g.index.duplicated(keep='first')]

# Calculate how many cells weren't found in solo_hashed out
def ret_htos(bcs, df_shan, b):
    # List of (barcode, HTO)
   barc_l = []
   #hto_l = []
   #samp_n = []
   ret_dict = {}
   doublet_n = 0
   negative_n = 0
   for bc in bcs:
       if bc in hash_data.obs_names:
           #hto_l.append(hash_data.obs[hash_data.obs_names == bc].Classification.values[0])
           hto_n = hash_data.obs[hash_data.obs_names == bc].Classification.values[0]
           #samp_n.append(df_shan[(df_shan["cDNA_ID"] == b) & (df_shan["hashtag" == hto_n]), "sample_ID"].values[0])
           if hto_n == 'Doublet':
               doublet_n += 1
        
           elif hto_n == 'Negative':
               negative_n += 1

           else:
               barc_l.append(bc)
               try:
                   ret_dict[bc] = df_shan.loc[(df_shan["cDNA_ID"] == b) & (df_shan["hashtag"] == hto_n), "SubID"].values[0]
               except:
                   if 'NPSAD-20201021-A1-cDNA' != b or 'NPSAD-20201021-A2-cDNA' != b or 'NPSAD-20201022-A1-cDNA' != b or 'NPSAD-20201022-A2-cDNA' != b:
                       print(f'For the Sample: {b} and for the HTO tag {hto_n} there was no data in the excel file!')


   return [barc_l, ret_dict, doublet_n, negative_n] #zip(barc_l, samp_n)



def ret_samp_names(y, bc_dict):
   return bc_dict[y]
   


adata = sc.read_10x_mtx(snakemake.input[1][:-13], make_unique=True, var_names= "gene_ids", cache=True)
solo_run_info = []
solo_run_info.append(( 'Started with cells', adata.n_obs))
solo_run_info.append(( 'Started with genes', adata.n_vars))
adata.var_names_make_unique()
adata.var["gene_id"] = adata.var.index.values
adata.var["gene_name"] = adata.var.gene_id.map(t2g["gene_name"])
adata.var_names = adata.var_names.to_series().map(lambda x: x + '_index')
adata = adata[:, pd.notna(adata.var["gene_name"])]
adata.var_names_make_unique()
adata.X = adata.X.astype('float64')
solo_run_info.append(( 'Unique genes', adata.n_vars))
   
# Filter data using cell level metrics
sc.pp.filter_cells(adata, min_genes=min_genes)
sc.pp.filter_genes(adata, min_cells=min_cells)
solo_run_info.append(( 'min #genes expressed per cell', min_genes))
solo_run_info.append(( 'min #cells expressing per gene', min_cells))
solo_run_info.append(( 'Retained cells after previous filter', adata.n_obs))
solo_run_info.append(( 'Retained genes after previous filter', adata.n_vars))
   
# Filter data wrt mito content
adata.var["mito"] = adata.var_names.str.startswith('MT-')
sc.pp.calculate_qc_metrics(adata, inplace=True, qc_vars=["mito"])
adata = adata[adata.obs["pct_counts_mito"]< max_mito, :]
solo_run_info.append(( 'max percent mito content per cell', max_mito))
solo_run_info.append(( 'cells with low mito percent', adata.n_obs))

# Download Shan's spreadsheet if already not present
try:
   wb = load_workbook('wet_lab_data_shan.xlsx')
except:
   get_file = subprocess.run(["wget", "--no-check-certificate", "https://docs.google.com/spreadsheets/d/10HYHDlsvtmBK_j3AgfPCnvaNohAg-OBhXi0iOzR9e0A/export?format=xlsx", "-O", "wet_lab_data_shan.xlsx"])
   wb = load_workbook('wet_lab_data_shan.xlsx')
   
df_l = []
for row in wb[wb.sheetnames[0]].iter_rows(min_col=2, max_col=47, min_row=1, values_only=True):
   df_l.append([re.sub('[-/\(\)]', '_', str(row[0])).replace(' ', ''), row[9], row[44].replace('#', 'HTO'), row[45]])


headers=df_l.pop(0)
headers[0] = headers[0].replace('.', '_')
headers[3] = 'barcode'
# Columns: Sub_ID, cDNA_ID, hashtag, barcode
df = pd.DataFrame(df_l, columns=headers)

# Demultiplex and assign samples (after downloading the spreadsheet)
hto_tags = ret_htos(adata.obs_names.tolist(), df, batch)
#barc_list = []
#samp_name_list = []

# hto_tags contain (in sequence): barcode, dictionary with barcodes as key and HTO# as value
#, no. of doublet cells, no. of negative cells]
adata = adata[hto_tags[0]].copy()
adata.obs['SubID'] = adata.obs_names.to_series().apply(ret_samp_names, args=(hto_tags[1],))

solo_run_info.append(('Doublets #cells', hto_tags[2]))
solo_run_info.append(('Negative #cells', hto_tags[3]))
adata.obs['batch'] = batch
adata.obs['round_num'] = r_num

solo_run_info.append(( 'After demultiplexing #cells', adata.n_obs))


#adata.obs['sample_name'] = ret_samp_names(adata.obs_names.tolist(), barc_list, HTO_list, df)


solo_run_df = pd.DataFrame(solo_run_info, columns=['Observations', 'Vals'])
solo_run_df.to_csv(snakemake.output[2], sep = "\t", index=False)
   
adata.write(snakemake.output[1])